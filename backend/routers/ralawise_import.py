"""
Ralawise importer — reads the Ralawise product spreadsheet (.xlsm/.xlsx) and
imports/updates products with their images, colours (with real RGB swatches) and
sizes, mirroring every image to our own R2 so nothing hotlinks to the supplier.

The Ralawise export is ONE ROW PER SKU (style × colour × size), ~23k rows. We
group by Style Code into ~600 products, each carrying its colour list (name + hex
from the RGB column + per-colour image) and size list.

Matching: we pass the Style Code as source_sku and reuse the same id-slug rule the
normal importer uses, so re-importing UPDATES existing products (fills in the
missing images/colours) rather than creating duplicates.

Endpoints:
  POST /admin/ralawise/preview  → parse an uploaded file, return a summary (no writes)
  POST /admin/ralawise/import   → parse + import/update + mirror images to R2
"""
from __future__ import annotations

import io
import re
from collections import OrderedDict
from typing import Dict, List, Optional

from fastapi import Depends, UploadFile, File, Form, HTTPException, BackgroundTasks
from pydantic import BaseModel

from deps import api_router, db, require_admin

# In-memory job registry for import progress (survives for the process lifetime,
# which is fine — a job completes in a few minutes).
_JOBS: Dict[str, Dict] = {}

# Column headers in the Ralawise export (0-indexed positions are resolved by
# name at runtime so a reordered export still works).
COL = {
    "style_code": "Style Code",
    "brand": "Brand",
    "style_name": "Style Name",
    "colour_name": "Colour Name",
    "size_name": "Size Name",
    "retail_desc": "Retail Description",
    "product_type": "Product Type",
    "categorisation": "Categorisation",
    "rgb": "RGB",
    "single_price": "Single Price",
    "pack_price": "Pack Price",
    "primary_image": "Primary Product Image URL",
    "colour_image": "Colour Image",
    "sku_status": "Sku Status",
}

# Ralawise "Product Type" / name hints → our catalogue category slugs.
_CATEGORY_HINTS = [
    ("hoodie", "hoodies"), ("hood", "hoodies"),
    ("sweat", "sweatshirts"), ("jumper", "sweatshirts"),
    ("polo", "polos"),
    ("t-shirt", "t-shirts"), ("tee", "t-shirts"), ("t shirt", "t-shirts"),
    ("jacket", "jackets"), ("gilet", "jackets"), ("bodywarmer", "jackets"),
    ("softshell", "jackets"), ("fleece", "jackets"),
    ("polo", "polos"),
    ("trouser", "bottoms"), ("jogger", "bottoms"), ("short", "bottoms"), ("legging", "bottoms"),
    ("cap", "hats"), ("hat", "hats"), ("beanie", "hats"),
    ("bag", "bags"), ("rucksack", "bags"), ("holdall", "bags"), ("tote", "bags"),
    ("apron", "aprons"),
    ("hi vis", "hi-vis"), ("hi-vis", "hi-vis"), ("hivis", "hi-vis"), ("high vis", "hi-vis"),
    ("vest", "t-shirts"), ("tank", "t-shirts"),
]


def _rgb_to_hex(rgb) -> Optional[str]:
    """'219 234 244' or '219,234,244' → '#dbeaf4'. Returns None if unparseable."""
    if not rgb:
        return None
    parts = re.split(r"[\s,]+", str(rgb).strip())
    try:
        nums = [max(0, min(255, int(float(p)))) for p in parts if p != ""][:3]
        if len(nums) != 3:
            return None
        return "#{:02x}{:02x}{:02x}".format(*nums)
    except Exception:
        return None


def _category_for(name: str, product_type: str, categorisation: str) -> str:
    hay = f"{name} {product_type} {categorisation}".lower()
    for needle, slug in _CATEGORY_HINTS:
        if needle in hay:
            return slug
    return "t-shirts"


# Preferred size display order.
_SIZE_ORDER = ["XXS", "XS", "S", "M", "L", "XL", "2XL", "XXL", "3XL", "XXXL", "4XL", "5XL", "6XL"]


def _sort_sizes(sizes: List[str]) -> List[str]:
    def key(s):
        u = s.upper().strip()
        return (_SIZE_ORDER.index(u) if u in _SIZE_ORDER else 99, u)
    return sorted(sizes, key=key)


def _parse_workbook(data: bytes) -> List[Dict]:
    """Parse the Ralawise xlsm/xlsx bytes into grouped product docs."""
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl isn't installed on the server — add 'openpyxl' to requirements.txt.")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {key: (header.index(col) if col in header else None) for key, col in COL.items()}
    if idx["style_code"] is None or idx["style_name"] is None:
        raise HTTPException(400, "This doesn't look like a Ralawise export (missing Style Code / Style Name columns).")

    def cell(row, key):
        i = idx.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    products: "OrderedDict[str, Dict]" = OrderedDict()
    for row in rows:
        style = cell(row, "style_code")
        name = cell(row, "style_name")
        if not style or not name:
            continue
        status = str(cell(row, "sku_status") or "").strip().lower()
        if status in ("discontinued", "deleted", "obsolete"):
            continue
        style = str(style).strip()
        p = products.get(style)
        if p is None:
            p = {
                "style_code": style,
                "name": str(name).strip(),
                "brand": str(cell(row, "brand") or "").strip(),
                "description": str(cell(row, "retail_desc") or "").strip(),
                "product_type": str(cell(row, "product_type") or "").strip(),
                "categorisation": str(cell(row, "categorisation") or "").strip(),
                "image": "",
                "_colours": OrderedDict(),   # name → {name, hex, image}
                "_sizes": set(),
                "_price": None,
            }
            products[style] = p

        # image (first non-empty wins for the product main image)
        if not p["image"]:
            img = cell(row, "primary_image")
            if img:
                p["image"] = str(img).strip()

        # price — take the lowest single price seen (the "from" price)
        price = cell(row, "single_price") or cell(row, "pack_price")
        try:
            price = float(price) if price is not None else None
        except Exception:
            price = None
        if price is not None and price > 0:
            p["_price"] = price if p["_price"] is None else min(p["_price"], price)

        # colour
        cname = cell(row, "colour_name")
        if cname:
            cname = str(cname).strip()
            if cname not in p["_colours"]:
                p["_colours"][cname] = {
                    "name": cname,
                    "hex": _rgb_to_hex(cell(row, "rgb")) or "",
                    "image": str(cell(row, "colour_image") or "").strip(),
                }

        # size
        sname = cell(row, "size_name")
        if sname:
            p["_sizes"].add(str(sname).strip())

    # finalise into import docs
    docs = []
    for style, p in products.items():
        colours = list(p["_colours"].values())
        docs.append({
            "source_sku": style,
            "name": p["name"],
            "brand": p["brand"],
            "description": p["description"],
            "price": round(p["_price"], 2) if p["_price"] else 0.0,
            "source_price": p["_price"],
            "image": p["image"],
            "category": _category_for(p["name"], p["product_type"], p["categorisation"]),
            "colors": colours,
            "sizes": _sort_sizes(list(p["_sizes"])),
            "source": "ralawise",
        })
    return docs


class RalawiseImportResult(BaseModel):
    ok: bool
    products: int
    with_images: int
    imported: int = 0
    updated: int = 0
    images_mirrored: int = 0
    images_failed: int = 0
    sample: List[Dict] = []
    error: Optional[str] = None


@api_router.post("/admin/ralawise/preview", dependencies=[Depends(require_admin)])
async def ralawise_preview(file: UploadFile = File(...)):
    """Parse the file and return a summary — no database writes."""
    data = await file.read()
    docs = _parse_workbook(data)
    with_images = sum(1 for d in docs if d["image"])
    sample = [{
        "style_code": d["source_sku"], "name": d["name"], "brand": d["brand"],
        "price": d["price"], "category": d["category"],
        "colours": len(d["colors"]), "sizes": len(d["sizes"]),
        "has_image": bool(d["image"]),
    } for d in docs[:12]]
    return RalawiseImportResult(ok=True, products=len(docs), with_images=with_images, sample=sample)


@api_router.post("/admin/ralawise/import", dependencies=[Depends(require_admin)])
async def ralawise_import(background_tasks: BackgroundTasks, file: UploadFile = File(...), mirror_images: bool = Form(True)):
    """Start a Ralawise import as a background JOB and return a job id straight
    away. The page then polls /admin/ralawise/status to show live progress and,
    if anything fails, an error that stays on screen. This avoids the request
    timing out on large files / hundreds of images."""
    from server import _slugify_source_sku
    import uuid as _uuid

    data = await file.read()

    # Parse up front so a bad file fails immediately (visible error), before we
    # start a job. This is fast (no network).
    try:
        docs = _parse_workbook(data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Couldn't read the file: {e}")

    for d in docs:
        d["id"] = _slugify_source_sku(d["name"], d["source_sku"])

    job_id = _uuid.uuid4().hex[:12]
    _JOBS[job_id] = {
        "id": job_id,
        "phase": "starting",
        "total": len(docs),
        "products_done": 0,
        "images_total": 0,
        "images_done": 0,
        "images_failed": 0,
        "imported": 0,
        "updated": 0,
        "error": None,
        "finished": False,
    }
    background_tasks.add_task(_run_ralawise_job, job_id, docs, bool(mirror_images))
    return {"ok": True, "job_id": job_id, "products": len(docs)}


@api_router.get("/admin/ralawise/status", dependencies=[Depends(require_admin)])
async def ralawise_status(job_id: str):
    job = _JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Unknown job id (it may have expired — start the import again).")
    return job


async def _run_ralawise_job(job_id: str, docs: List[Dict], mirror_images: bool) -> None:
    """The whole import, run in the background with live progress in _JOBS."""
    from server import PRODUCTS, _apply_imported_product, _mirror_external_image
    from datetime import datetime, timezone
    import asyncio

    job = _JOBS[job_id]
    try:
        now = datetime.now(timezone.utc).isoformat()
        existing_ids = set(PRODUCTS.keys())
        job["imported"] = sum(1 for d in docs if d["id"] not in existing_ids)
        job["updated"] = sum(1 for d in docs if d["id"] in existing_ids)

        # ---- Phase 1: write all products (fast) ----
        job["phase"] = "importing products"
        for i, d in enumerate(docs):
            doc = {
                "id": d["id"], "source_sku": d["source_sku"], "name": d["name"],
                "brand": d["brand"], "description": d["description"], "price": d["price"],
                "source_price": d["source_price"], "image": d["image"], "category": d["category"],
                "colors": d["colors"], "sizes": d["sizes"], "source": "ralawise",
                "active": True, "imported_at": now,
            }
            await db.imported_products.update_one({"id": d["id"]}, {"$set": doc}, upsert=True)
            _apply_imported_product(doc)
            job["products_done"] = i + 1

        # ---- Phase 2: mirror images to R2 (the slow part) ----
        if mirror_images:
            job["phase"] = "copying images to storage"
            # collect distinct urls
            url_to_docs: Dict[str, list] = {}
            for d in docs:
                for key, holder in [("main", d)] + [("colour", c) for c in d["colors"]]:
                    u = holder.get("image")
                    if u and "r2.dev" not in str(u) and not str(u).startswith("data:"):
                        url_to_docs.setdefault(u, []).append(holder)
            job["images_total"] = len(url_to_docs)

            sem = asyncio.Semaphore(10)
            done = 0
            lock = asyncio.Lock()

            async def do_one(u, holders):
                nonlocal done
                async with sem:
                    try:
                        mirrored = await _mirror_external_image(u)
                    except Exception:
                        mirrored = None
                async with lock:
                    done += 1
                    job["images_done"] = done
                    if mirrored:
                        for h in holders:
                            h["image"] = mirrored
                    else:
                        job["images_failed"] += 1

            # process in batches so progress ticks up smoothly + we can persist
            items = list(url_to_docs.items())
            BATCH = 40
            for start in range(0, len(items), BATCH):
                chunk = items[start:start + BATCH]
                await asyncio.gather(*[do_one(u, h) for u, h in chunk])
                # persist the docs touched in this batch
                touched = {h_id for _, hs in chunk for h_id in [id(h) for h in hs]}
                # re-save every product (cheap enough; ensures colour image urls persist)
            # persist all products' updated image + colours
            for d in docs:
                try:
                    await db.imported_products.update_one(
                        {"id": d["id"]},
                        {"$set": {"image": d["image"], "colors": d["colors"]}}
                    )
                    p = PRODUCTS.get(d["id"])
                    if p:
                        p["image"] = d["image"]
                        p["colors"] = d["colors"]
                except Exception:
                    pass

        job["phase"] = "done"
        job["finished"] = True
    except Exception as e:
        job["error"] = str(e)[:400]
        job["phase"] = "failed"
        job["finished"] = True
